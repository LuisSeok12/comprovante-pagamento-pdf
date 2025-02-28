#!/usr/bin/env python3
"""
Script para processar PDFs de comprovantes de pagamento e atualizar um arquivo Excel.

Este script realiza as seguintes etapas:
1. Permite ao usuário selecionar um arquivo Excel com os dados das transações.
2. Permite ao usuário escolher a pasta base onde os PDFs estão organizados (por ano e mês).
3. Para cada registro do Excel, o script busca o PDF correspondente (com base na data) e procura
   por uma transação específica (usando o número da fatura ou o valor do pagamento).
4. Se a transação for encontrada, a página é extraída e salva em uma subpasta "Notas".
5. Atualiza a coluna "Encontrado" no arquivo Excel para indicar se o comprovante foi localizado.

*Observação:* A senha para desbloqueio dos PDFs foi removida, pois os arquivos não estão criptografados.
"""

import os
import re
import pandas as pd
from PyPDF2 import PdfReader, PdfWriter
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def escolher_arquivo_excel():
    """
    Abre uma janela para o usuário selecionar o arquivo Excel.

    Retorna:
        str: Caminho do arquivo Excel selecionado.
    """
    root = tk.Tk()
    root.withdraw()  # Oculta a janela principal
    file_path = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
    )
    root.destroy()
    return file_path

def escolher_pasta_base():
    """
    Abre uma janela para o usuário selecionar a pasta base dos PDFs.

    Retorna:
        str: Caminho da pasta base selecionada.
    """
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(title="Selecione a pasta base dos PDFs")
    root.destroy()
    return folder_path

def formatar_valor_ptbr(valor_float: float) -> str:
    """
    Converte um valor float no formato monetário brasileiro.
    Exemplo: 170606.16 -> 'R$ 170.606,16'

    Args:
        valor_float (float): Valor a ser formatado.

    Retorna:
        str: Valor formatado.
    """
    valor_str_us = f"{valor_float:,.2f}"
    valor_str_br = valor_str_us.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {valor_str_br}"

def processar_registro(row, base_pdf_folder):
    """
    Processa um registro do DataFrame, buscando a transação no PDF correspondente
    e exportando a página onde a transação foi encontrada.

    Args:
        row (pandas.Series): Linha do DataFrame com os dados da transação.
        base_pdf_folder (str): Caminho da pasta base onde estão os PDFs.

    Retorna:
        bool: True se a transação for encontrada e processada com sucesso, False caso contrário.
    """
    # Extrai data e valor da transação
    vencimento = row["Vencimento"]
    valor_pagamento = row["( R$ )"]

    try:
        data_formatada = pd.to_datetime(vencimento)
    except Exception as e:
        print(f"[ERRO] Não foi possível converter a data '{vencimento}': {e}")
        return False

    # Define as pastas com base na data: ano e mês
    ano = data_formatada.year
    pasta_ano = f"Comprovantes de pagamento - {ano}"
    month_folder = data_formatada.strftime("%m.%Y")
    pasta_pdf_mes = os.path.join(base_pdf_folder, pasta_ano, month_folder)
    
    if not os.path.exists(pasta_pdf_mes):
        print(f"[ERRO] Pasta do mês '{pasta_pdf_mes}' não encontrada.")
        return False

    # Define o nome do PDF com base no dia e mês (ex.: "15 03.pdf")
    nome_pdf = data_formatada.strftime("%d %m") + ".pdf"
    caminho_pdf = os.path.join(pasta_pdf_mes, nome_pdf)

    if not os.path.exists(caminho_pdf):
        print(f"[ERRO] Arquivo PDF '{caminho_pdf}' não encontrado.")
        return False

    try:
        leitor = PdfReader(caminho_pdf)
    except Exception as e:
        print(f"[ERRO] Erro ao abrir o PDF '{caminho_pdf}': {e}")
        return False

    # Define o critério de busca: prioriza "Número da Fatura", se disponível; caso contrário, utiliza o valor do pagamento
    numero_fatura = row.get("Número da Fatura", None)
    if numero_fatura is not None and pd.notna(numero_fatura):
        search_value = str(numero_fatura).strip()
        search_type = "Número da Fatura"
    else:
        try:
            valor_float = float(valor_pagamento)
        except Exception as e:
            print(f"[ERRO] Valor de pagamento inválido '{valor_pagamento}': {e}")
            return False
        valor_str = f"{valor_float:.2f}"
        search_value = formatar_valor_ptbr(valor_float)
        search_type = "Valor pagamento"

    # Cria um padrão regex para procurar o valor ou número da fatura no texto do PDF
    search_value_esc = re.escape(search_value)
    search_value_pattern = search_value_esc.replace(r'\ ', r'\s*')
    pattern = re.compile(search_value_pattern)

    pagina_alvo = None
    for indice, pagina in enumerate(leitor.pages):
        try:
            texto = pagina.extract_text()
            if texto:
                texto = texto.replace("\n", " ")
        except Exception as e:
            print(f"[ERRO] Falha ao extrair texto da página {indice} do PDF '{caminho_pdf}': {e}")
            continue

        if texto and re.search(pattern, texto):
            pagina_alvo = indice
            print(f"[INFO] Transação com {search_type} '{search_value}' encontrada na página {indice} do PDF '{nome_pdf}'.")
            break

    if pagina_alvo is None:
        print(f"[AVISO] Transação com {search_type} '{search_value}' não encontrada em '{nome_pdf}'.")
        return False

    # Cria o PDF contendo apenas a página com a transação encontrada
    writer = PdfWriter()
    writer.add_page(leitor.pages[pagina_alvo])

    # Define o nome do arquivo de saída conforme o critério de busca
    if search_type == "Número da Fatura":
        nome_saida = f"Comprov_nf {search_value}_{data_formatada.strftime('%d %m')}.pdf"
    else:
        nome_saida = f"Comprov_nf {valor_str}_{data_formatada.strftime('%d %m')}.pdf"

    # Cria a pasta "Notas" dentro da pasta do mês (caso não exista)
    pasta_saida = os.path.join(pasta_pdf_mes, "Notas")
    if not os.path.exists(pasta_saida):
        os.makedirs(pasta_saida)
    caminho_saida = os.path.join(pasta_saida, nome_saida)

    try:
        with open(caminho_saida, "wb") as f_out:
            writer.write(f_out)
        print(f"[SUCESSO] Página exportada para '{caminho_saida}'.")
        return True
    except Exception as e:
        print(f"[ERRO] Falha ao exportar o PDF '{caminho_saida}': {e}")
        return False

def main():
    """
    Função principal que orquestra o fluxo do script:
      1. Seleciona o arquivo Excel.
      2. Seleciona a pasta base dos PDFs.
      3. Processa cada registro do Excel.
      4. Atualiza a coluna "Encontrado" no arquivo Excel.
    """
    arquivo_excel = escolher_arquivo_excel()
    if not arquivo_excel:
        print("Nenhum arquivo foi selecionado. Encerrando o programa.")
        return

    base_pdf_folder = escolher_pasta_base()
    if not base_pdf_folder:
        print("Nenhuma pasta foi selecionada. Encerrando o programa.")
        return

    try:
        df = pd.read_excel(arquivo_excel)
        # Remove espaços extras dos nomes das colunas
        df.columns = df.columns.str.strip()
        expected_cols = ["Vencimento", "( R$ )", "Encontrado"]
        if not all(col in df.columns for col in expected_cols):
            print("Os nomes das colunas não estão como esperado. Colunas encontradas:", df.columns.tolist())
            if "Data de\nPagamento" in df.columns:
                df.rename(columns={"Data de\nPagamento": "Vencimento"}, inplace=True)
            if "Valor\npagamento\nlíquido (R$)" in df.columns:
                df.rename(columns={"Valor\npagamento\nlíquido (R$)": "( R$ )"}, inplace=True)
        print("Colunas utilizadas:", df.columns.tolist())
    except Exception as e:
        print(f"[ERRO] Não foi possível ler o arquivo Excel '{arquivo_excel}': {e}")
        return

    # Processa cada registro e atualiza a coluna "Encontrado"
    for i, row in df.iterrows():
        try:
            print(f"\n[INFO] Processando registro {i+1} - Vencimento: {row['Vencimento']} | ( R$ ): {row['( R$ )']}")
        except KeyError as e:
            print(f"[ERRO] Chave não encontrada: {e}")
            continue
        resultado = processar_registro(row, base_pdf_folder)
        df.at[i, "Encontrado"] = "Sim" if resultado else "Não"

    # Atualiza a coluna "Encontrado" no arquivo Excel mantendo o restante da formatação
    try:
        wb = load_workbook(arquivo_excel)
        ws = wb.active
        encontrado_col = None
        for cell in ws[1]:
            if cell.value == "Encontrado":
                encontrado_col = cell.column_letter
                break

        if encontrado_col is None:
            print("[ERRO] Coluna 'Encontrado' não encontrada no arquivo Excel.")
        else:
            for i, row in df.iterrows():
                ws[f"{encontrado_col}{i+2}"].value = row["Encontrado"]
            wb.save(arquivo_excel)
            print(f"\n[INFO] Coluna 'Encontrado' atualizada no arquivo '{arquivo_excel}'.")
    except Exception as e:
        print(f"[ERRO] Falha ao atualizar o arquivo Excel: {e}")

if __name__ == "__main__":
    main()
